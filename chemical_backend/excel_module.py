import os
from openpyxl import load_workbook, Workbook

def save_to_excel(weight, image_path, current_time, ocr_text, person_id,
                  excel_file='measurement_log.xlsx'):
    """
    保存数据到Excel，包含重量、OCR文字、操作人ID，以及图片路径（文本形式）
    """
    # 准备新行数据（与列标题顺序对应）
    new_row_data = [
        current_time.strftime('%Y-%m-%d %H:%M:%S'),   # 称量时间
        weight,                                      # 重量(g)
        image_path,                                  # 图片路径（文本）
        os.path.basename(image_path) if image_path else None,  # 图片文件名
        ocr_text,                                    # OCR识别结果
        person_id,                                   # 操作人ID
        current_time.timestamp()                     # 记录时间戳
    ]
    headers = ['称量时间', '重量(g)', '图片路径', '图片文件名', 'OCR识别结果', '操作人ID', '记录时间戳']

    # 判断文件是否存在
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '称量记录'
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        next_row = 2

    # 写入文本数据（所有列）
    for col_idx, value in enumerate(new_row_data, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    # 不再插入图片，仅保存路径文本（已在第3列）
    # 如果希望保留行高/列宽等可以自行添加，但非必要

    wb.save(excel_file)
    print(f"数据（含图片路径）已写入 {excel_file}")
    return True